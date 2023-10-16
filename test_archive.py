import os
import zipfile
import xlrd
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader

from utils import ARCHIVE_PATH, TXT_PATH, PDF_PATH, XLS_PATH, XLSX_PATH

def test_pdf(download_pdf):
    pdf_size_real = os.path.getsize(PDF_PATH)
    pdf_name_real = os.path.basename(PDF_PATH)
    reader = PdfReader(PDF_PATH)
    pdf_pages_real = len(reader.pages)
    pdf_text_real = reader.pages[0].extract_text()
    with zipfile.ZipFile(file=ARCHIVE_PATH, mode='a') as z:
        z.write(filename=PDF_PATH, arcname='sample.pdf')
    with zipfile.ZipFile(file=ARCHIVE_PATH, mode='r') as z:
        pdf_file_zip = PdfReader(z.open('sample.pdf', 'r'))
        pages_pdf_zip = len(pdf_file_zip.pages)
        text_pdf_zip = pdf_file_zip.pages[0].extract_text()
        pdf_info = z.getinfo('sample.pdf')
        pdf_name_zip = pdf_info.filename
        pdf_size_zip = pdf_info.file_size
        assert pdf_name_zip == pdf_name_real
        assert pdf_size_zip == pdf_size_real
        assert pages_pdf_zip == pdf_pages_real
        assert pdf_text_real == text_pdf_zip



def test_txt(download_txt):
    txt_size_real = os.path.getsize(TXT_PATH)
    txt_name_real = os.path.basename(TXT_PATH)
    with zipfile.ZipFile(file=ARCHIVE_PATH, mode='a') as z:
        z.write(filename=TXT_PATH, arcname='sample.txt')
    with zipfile.ZipFile(file=ARCHIVE_PATH, mode='r') as z:
        txt_info_zip = z.getinfo('sample.txt')
        txt_name_zip = txt_info_zip.filename
        txt_size_zip = txt_info_zip.file_size
        with z.open('sample.txt') as txt_zip:
            assert 'Lorem ipsum dolor sit amet, consectetur adipiscing elit.' in txt_zip.read().decode('utf-8')
    assert txt_name_zip == txt_name_real
    assert txt_size_zip == txt_size_real

def test_xls(download_xls):
    xls_file_real = xlrd.open_workbook(XLS_PATH)
    xls_size_real = os.path.getsize(XLS_PATH)
    xls_name_real = os.path.basename(XLS_PATH)
    with zipfile.ZipFile(file=ARCHIVE_PATH, mode='a') as z:
        z.write(filename=XLS_PATH, arcname='sample.xls')
    with zipfile.ZipFile(file=ARCHIVE_PATH, mode='r') as z:
        with z.open('sample.xls') as book:
            xls_file_zip = xlrd.open_workbook(file_contents=book.read())
        xls_info_zip = z.getinfo('sample.xls')
        xls_name_zip = xls_info_zip.filename
        xls_size_zip = xls_info_zip.file_size
    assert xls_name_zip == xls_name_real
    assert xls_size_zip == xls_size_real
    assert xls_file_real.sheet_names() == xls_file_zip.sheet_names()

def test_xlsx(download_xlsx):
    xlsx_text_real = load_workbook(XLSX_PATH).active.cell(row=2, column=1).value
    xlsx_size_real = os.path.getsize(XLSX_PATH)
    xlsx_name_real = os.path.basename(XLSX_PATH)
    with zipfile.ZipFile(file=ARCHIVE_PATH, mode='a') as z:
        z.write(filename=XLSX_PATH, arcname='sample.xlsx')
    with zipfile.ZipFile(file=ARCHIVE_PATH, mode='r') as z:
        with z.open('sample.xlsx') as book:
            xlsx_text_zip = load_workbook(book).active.cell(row=2, column=1).value
        xlsx_info = z.getinfo('sample.xlsx')
        xlsx_name = xlsx_info.filename
        xlsx_size = xlsx_info.file_size
    assert xlsx_name == xlsx_name_real
    assert xlsx_size == xlsx_size_real
    assert xlsx_text_real == xlsx_text_zip

