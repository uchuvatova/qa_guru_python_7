import os.path
from utils import RESOURCES_PATH, PROJECT_ROOT_PATH
import requests
from selene import query
from selene.support.shared import browser
from pypdf import PdfReader
from xlrd import open_workbook
from openpyxl import load_workbook
import zipfile, os
import shutil

def test_download_pdf_file_with_selene_by_href():
    browser.open("/pdf")
    href = browser.element("[href='/samples/document/pdf/sample2.pdf']").get(query.attribute("href"))
    content = requests.get(href).content
    with open("resources/sample.pdf", 'wb') as f:
        f.write(content)
    reader = PdfReader(os.path.join(RESOURCES_PATH, "sample.pdf"))
    number_of_pages = len(reader.pages)
    print(number_of_pages)
    page = reader.pages[0]
    text = page.extract_text()
    assert "This is an example of a user fillable PDF form." in text

def test_download_txt_file_with_selene_by_href():
    browser.open("/txt")
    href = browser.element("[href='/samples/document/txt/sample1.txt']").get(query.attribute("href"))
    content = requests.get(href).content
    with open("resources/sample.txt", 'wb') as f:
        f.write(content)
    with open('resources/sample.txt') as f:
        text = f.read()
        assert "Lorem ipsum dolor sit amet, consectetur adipiscing elit." in text

def test_download_xls_file_with_selene_by_href():
    browser.open("/xls")
    href = browser.element("[href='/samples/document/xls/sample1.xls']").get(query.attribute("href"))
    content = requests.get(href).content
    with open("resources/sample.xls", 'wb') as f:
        f.write(content)
    book = open_workbook("resources/sample.xls")
    assert book.sheet_names() == ['Example Test', 'Format Abbr.', 'Readme']
    count_sheets = f"{book.nsheets}"
    assert count_sheets == "3"
    sheet = book.sheet_by_index(0)
    text = sheet.cell_value(1, 1)
    assert "What C datatypes are 8 bits?" in text

def test_download_xlsx_file_with_selene_by_href():
    browser.open("/xlsx")
    href = browser.element("[href='/samples/document/xlsx/sample1.xlsx']").get(query.attribute("href"))
    content = requests.get(href).content
    with open("resources/sample.xlsx", 'wb') as f:
        f.write(content)
    workbook = load_workbook('resources/sample.xlsx')
    sheet = workbook.active
    text = sheet.cell(row=2, column=1).value
    assert "2121" in text

def test_create_zip():
    shutil.make_archive('test_archive', 'zip', './resources', './tmp')


    #path = "/resources/"
    #file_dir = os.listdir(path)
#
    #with zipfile.ZipFile('test.zip', mode='w', \
    #                     compression=zipfile.ZIP_DEFLATED) as zf:
    #    for file in file_dir:
    #        add_file = os.path.join(path, file)
    #        zf.write(add_file)
    #with zipfile.ZipFile('test.zip', mode='a') as zf:
    #    for file in zf.namelist():
    #        print(file)
    #with zipfile.ZipFile('test.zip', mode='a') as zf:
    #    for file in zf.infolist():
    #        # имя файла в архиве без пути
    #        name = os.path.basename(file.filename)
    #        # печатаем имя, начальный размер,
    #        # размер в архиве, дата файла
    #        print(f"{name},\t{file.file_size},\t{file.compress_size}")